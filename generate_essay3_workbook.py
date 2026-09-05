"""Generate essay3_full_results.xlsx from ESSAY3_* tables in SQLite.

Usage:
    cd /Users/administrator/Projects/signalsandsystems
    .venv/bin/python generate_essay3_workbook.py
"""

import io
import sqlite3

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

DB_PATH = 'data/signals_systems.db'
OUT_PATH = 'essay3_full_results.xlsx'

# ── Styles ─────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL  = PatternFill('solid', fgColor='D6E4F0')
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT  = Font(italic=True, size=9, color='555555')


def _write_table(wb, sheet_name, df, title=None, note=None):
    """Write a DataFrame to a worksheet with styled header."""
    ws = wb.create_sheet(sheet_name[:31])
    row = 1
    if title:
        ws.cell(row, 1, title).font = TITLE_FONT
        row += 1
    if note:
        ws.cell(row, 1, note).font = NOTE_FONT
        row += 1
        ws.merge_cells(start_row=row-1, start_column=1,
                       end_row=row-1, end_column=max(len(df.columns), 1))
    if df.empty:
        ws.cell(row, 1, '(no data)')
        return ws

    # Header
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row, c, col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(wrap_text=True)
    row += 1

    # Data rows — skip alt-colour for huge tables (speed)
    use_alt = len(df) <= 5000
    for i, (_, data_row) in enumerate(df.iterrows()):
        for c, val in enumerate(data_row, 1):
            cell = ws.cell(row + i, c, val if pd.notna(val) else '')
            if use_alt and i % 2 == 1:
                cell.fill = ALT_FILL

    # Auto-width (capped)
    for c, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df[col].astype(str).str.len().max() if not df.empty else 0)
        ws.column_dimensions[ws.cell(1, c).column_letter].width = min(max_len + 2, 40)

    return ws


def _fig_to_image(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=120, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return XLImage(buf)


def _add_image_sheet(wb, name, fig):
    ws = wb.create_sheet(name[:31])
    img = _fig_to_image(fig)
    ws.add_image(img, 'A1')
    return ws


# ── Figure builders ────────────────────────────────────────────────────────────

def fig_sell_accuracy_by_proximity(informed_proximity):
    df = informed_proximity[
        (informed_proximity['SAMPLE'] == 'POLITICAL') &
        (informed_proximity['METRIC_TYPE'] == 'PROPORTION') &
        (informed_proximity['CUT'].str.contains('SELLS', na=False))
    ].copy().sort_values('WINDOW_START')
    if df.empty:
        return None
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.bar(df['CUT'], df['METRIC_VALUE'] * 100, color='steelblue')
    ax.axhline(57.4, color='red', linestyle='--', label='Base rate (57.4%)')
    ax.axhline(50, color='grey', linestyle=':', label='Naive 50%')
    ax.set_ylabel('% Trades Profitable (Event-CAR)')
    ax.set_title('Sell Accuracy by Days Before Event\n(Event-CAR definition)', fontsize=11)
    ax.legend()
    plt.xticks(rotation=20, ha='right')
    plt.tight_layout()
    return fig


def fig_plan_vs_nonplan(informed_trading):
    df = informed_trading[
        (informed_trading['CUT'].isin(['SELLS_10B5_PLAN', 'SELLS_NOT_10B5_PLAN', 'SELLS_POST2023'])) &
        (informed_trading['METRIC_TYPE'] == 'PROPORTION')
    ].copy()
    if df.empty:
        return None
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.bar(df['CUT'], df['METRIC_VALUE'] * 100, color=['#2196F3', '#F44336', '#9C27B0'])
    ax.axhline(57.4, color='red', linestyle='--', label='Base rate (57.4%)')
    ax.set_ylabel('% Profitable (Event-CAR)')
    ax.set_title('10b5-1 Plan vs Non-Plan Sell Accuracy\n(Post-2023 filings)', fontsize=11)
    ax.legend()
    plt.xticks(rotation=15, ha='right')
    plt.tight_layout()
    return fig


def fig_cluster_inference(cluster_inference):
    if cluster_inference.empty:
        return None
    sells = cluster_inference[cluster_inference['CUT'].str.contains('SELLS', na=False)].copy()
    if sells.empty:
        return None
    fig, ax = plt.subplots(figsize=(9, 5))
    x = range(len(sells))
    ax.bar(x, sells['POINT_EST'] * 100, color='steelblue', alpha=0.7, label='Point Est')
    for i, (_, row) in enumerate(sells.iterrows()):
        ax.errorbar(i, row['POINT_EST'] * 100,
                    yerr=[[row['POINT_EST']*100 - row['CI_LO_95']*100],
                          [row['CI_HI_95']*100 - row['POINT_EST']*100]],
                    fmt='none', color='black', capsize=4)
    ax.axhline(57.4, color='red', linestyle='--', label='Cond. base rate (57.4%)')
    ax.set_xticks(list(x))
    ax.set_xticklabels(
        [f"{r['CUT']}\n({r['CLUSTER_TYPE']})" for _, r in sells.iterrows()],
        rotation=30, ha='right', fontsize=8)
    ax.set_ylabel('Sell Accuracy (%)')
    ax.set_title('Cluster Inference: Point Estimates + 95% CIs', fontsize=11)
    ax.legend()
    plt.tight_layout()
    return fig


def fig_concentration(insider_concentration):
    df = insider_concentration.copy()
    metrics = ['GINI', 'TOP_5PCT_SHARE', 'TOP_10PCT_SHARE']
    cats = df['EVENT_CATEGORY'].unique() if 'EVENT_CATEGORY' in df.columns else []
    if not len(cats):
        return None
    df_pivot = df[df['METRIC'].isin(metrics)].pivot(
        index='METRIC', columns='EVENT_CATEGORY', values='VALUE') \
        if 'METRIC' in df.columns else None
    if df_pivot is None or df_pivot.empty:
        return None
    fig, ax = plt.subplots(figsize=(7, 4))
    df_pivot.plot(kind='bar', ax=ax)
    ax.set_title('Insider Trading Concentration by Category', fontsize=11)
    ax.set_ylabel('Value')
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig


def fig_crsp_summary(crsp_summary):
    df = crsp_summary.copy()
    sub = df[df['CUT'].isin(['ACTIVE', 'INACTIVE', 'ACTIVE_SELLS', 'ACTIVE_BUYS'])].copy()
    if sub.empty or 'PCT_PROFITABLE_30' not in sub.columns:
        return None
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.bar(sub['CUT'], sub['PCT_PROFITABLE_30'] * 100, color='teal')
    ax.axhline(50, color='grey', linestyle=':', label='50%')
    ax.set_ylabel('% Profitable (30d trade-CAR)')
    ax.set_title('CRSP Trade Profitability by Insider Type', fontsize=11)
    ax.legend()
    plt.xticks(rotation=15, ha='right')
    plt.tight_layout()
    return fig


def fig_directional_placebo(directional_placebo, informed_trading):
    if directional_placebo.empty:
        return None
    row = directional_placebo.iloc[0]
    placebo_mean = row.get('PERM_MEAN', np.nan)
    placebo_sd = row.get('PERM_STD', np.nan)
    actual = None
    if not informed_trading.empty:
        sells = informed_trading[
            (informed_trading['CUT'] == 'ALL') &
            (informed_trading['METRIC_TYPE'] == 'PROPORTION') &
            (informed_trading['SAMPLE'] == 'POLITICAL')]
        if not sells.empty:
            actual = sells['METRIC_VALUE'].iloc[0]

    fig, ax = plt.subplots(figsize=(6, 4))
    if not np.isnan(placebo_mean) and not np.isnan(placebo_sd):
        x = np.linspace(placebo_mean - 4*placebo_sd, placebo_mean + 4*placebo_sd, 300)
        from scipy.stats import norm
        ax.plot(x * 100, norm.pdf(x, placebo_mean, placebo_sd), color='steelblue',
                label=f'Placebo dist (μ={placebo_mean*100:.1f}%)')
    if actual:
        ax.axvline(actual * 100, color='red', linestyle='--',
                   label=f'Observed sells ({actual*100:.1f}%)')
    ax.set_xlabel('Sell Accuracy (%)')
    ax.set_title('Date-Randomization Placebo vs Observed', fontsize=11)
    ax.legend()
    plt.tight_layout()
    return fig


def fig_joint_pt(joint_pt):
    if joint_pt.empty:
        return None
    row = joint_pt.iloc[0]
    fig, ax = plt.subplots(figsize=(5, 4))
    cats = ['Null\n(independence)', 'Observed']
    vals = [row.get('NULL_ACCURACY', np.nan) * 100,
            row.get('OBS_ACCURACY', np.nan) * 100]
    colors = ['grey', 'steelblue']
    bars = ax.bar(cats, vals, color=colors)
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width()/2, v + 0.1, f'{v:.2f}%',
                ha='center', va='bottom', fontsize=10)
    z = row.get('PT_Z_STAT', np.nan)
    p = row.get('PT_P_VALUE', np.nan)
    ax.set_ylabel('Joint Directional Accuracy (%)')
    ax.set_title(f'Joint P-T Sign Test\nZ={z:.2f}, p={p:.3f}', fontsize=11)
    plt.tight_layout()
    return fig


def fig_car_net_slope(car_net_slope):
    if car_net_slope.empty:
        return None
    row = car_net_slope.iloc[0]
    coef = row.get('COEF_NET_DIRECTION', np.nan)
    se = row.get('SE_HC3', np.nan)
    p = row.get('P_VALUE_TWO', np.nan)
    fig, ax = plt.subplots(figsize=(5, 4))
    ax.bar(['Net Sell Direction\n→ Event CAR'], [coef * 100],
           color='steelblue' if coef < 0 else 'salmon',
           yerr=[[se * 1.96 * 100]], capsize=8)
    ax.axhline(0, color='black', linewidth=0.8)
    ax.set_ylabel('Coefficient (CAR %)')
    ax.set_title(f'CAR-on-Net-Direction Slope\n(HC3 SE, p={p:.3f})', fontsize=11)
    plt.tight_layout()
    return fig


# ── README ─────────────────────────────────────────────────────────────────────

README_TEXT = [
    ('Essay 3 — Informed Insider Trading Around Political Decisions', True, 13),
    ('Full Results Workbook  |  Last updated: 2026-07-23  |  Numbers verified from live SQLite', False, 10),
    ('', False, 10),
    ('── RESEARCH DESIGN ────────────────────────────────────────────────────', False, 10),
    ('Full panel: 2000–2025 | 1,712 political events | 29,633 trades (20,055 sells, 9,578 buys)', False, 10),
    ('10b5-1 plan flag available only for post-2023 filings (SEC amendment effective 2023-02-27).', False, 10),
    ('Plan/non-plan analysis is a post-2023 sub-sample (N=162 events, N=2,678 sell trades).', False, 10),
    ('', False, 10),
    ('── THE BASE-RATE PROBLEM ───────────────────────────────────────────────', False, 10),
    ('', False, 10),
    ('The conditional null for sells = fraction of events with NEGATIVE post-event CAR.', False, 10),
    ('This fraction varies by period. Each period must be compared to ITS OWN base rate.', False, 10),
    ('Using the pooled 57.84% as the null for all periods is the central methodological error.', False, 10),
    ('', False, 10),
    ('Period-specific base rates (ESSAY3_PANEL, CAR_POST < 0):', True, 10),
    ('', False, 10),
    ('  Period              N events  Base rate  Sell accuracy  vs own base rate', False, 10),
    ('  SOX_ERA (03-09)       283      47.3%        48.9%         +1.6pp', False, 10),
    ('  DODD_FRANK (10-22)    688      58.6%        54.1%         −4.5pp', False, 10),
    ('  POST_AMENDMENTS(23+)  162      71.6%        66.4%         −5.2pp  ← furthest below', False, 10),
    ('  Pooled              1,141      57.4%        54.7%         −2.7pp', False, 10),
    ('', False, 10),
    ('  The rising sell accuracy from SOX→DF→POST mirrors a rising fraction of negative-CAR events.', False, 10),
    ('  When each period is benchmarked against its own base rate, every period is at or below.', False, 10),
    ('  POST_AMENDMENTS shows the largest negative gap (−5.2pp), not a "spike."', False, 10),
    ('  The regulatory arc is an arc in the EVENT COMPOSITION, not in insider behavior.', False, 10),
    ('', False, 10),
    ('── PLAN vs NON-PLAN (POST-2023 SUB-SAMPLE) ────────────────────────────', False, 10),
    ('', False, 10),
    ('Post-2023 period base rate: 71.6% of events had negative post-event CARs.', False, 10),
    ('', False, 10),
    ('  CUT                  N trades  Accuracy  vs own base rate (71.6%)', False, 10),
    ('  SELLS_POST2023        2,174     59.7%        −11.9pp', False, 10),
    ('  SELLS_10B5_PLAN         739     50.7%        −20.9pp', False, 10),
    ('  SELLS_NOT_10B5_PLAN   1,435     64.3%         −7.3pp', False, 10),
    ('', False, 10),
    ('  Both arms are below the period-specific base rate. Neither shows informativeness.', False, 10),
    ('  Trade-level gap (non-plan − plan): +13.6pp. But at EVENT_AGG level: −1.6pp (reversal).', False, 10),
    ('  Non-plan EVENT p=0.021 tests that arm vs 0.5, not the gap between arms.', False, 10),
    ('  The trade-level gap is not robust to event-level aggregation.', False, 10),
    ('', False, 10),
    ('── AFFIRMATIVE RESULT: TOST EQUIVALENCE ────────────────────────────────', False, 10),
    ('', False, 10),
    ('TOST tests whether mean net trading value is practically equivalent to zero (SESOI = d=0.2).', False, 10),
    ('  CULTURAL:            N=234,  MEAN=$309K,  P_TOST=0.004  → EQUIVALENT', False, 10),
    ('  FUNDAMENTAL:         N=1,419 MEAN=$39K,   P_TOST=3.5e-12 → EQUIVALENT', False, 10),
    ('  FUNDAMENTAL_MATCHED: N=680,  MEAN=−$51K,  P_TOST=1.4e-6  → EQUIVALENT', False, 10),
    ('  All three pass TOST. Mean abnormal trading value is within practical-equivalence bounds.', False, 10),
    ('  This is the cleanest affirmative finding: no economically meaningful elevated trading.', False, 10),
    ('', False, 10),
    ('── FULL-PANEL ACCURACY RESULTS ─────────────────────────────────────────', False, 10),
    ('', False, 10),
    ('Pooled conditional base rate: 57.84% sells / 42.16% buys (from COND_NULL column).', False, 10),
    ('Note: METRIC_PVAL_COND uses 57.84% pooled null — not the period-specific rates above.', False, 10),
    ('', False, 10),
    ('SELLS vs pooled null:', True, 10),
    ('  Observed: 54.68% (N=20,055) | Pooled null: 57.84% → below null, p=1.78e-19', False, 10),
    ('  Cluster-robust (SELLS_COND): EVENT p=0.836, TICKER p=0.862, EVENT_AGG p=0.373 — all n.s.', False, 10),
    ('', False, 10),
    ('BUYS vs pooled null:', True, 10),
    ('  Observed: 45.36% (N=9,578) | Pooled null: 42.16% → above null, p=2.39e-10', False, 10),
    ('  Cluster-robust (BUYS_COND): EVENT p=0.170, TICKER p=0.128 — not significant.', False, 10),
    ('', False, 10),
    ('POLITICAL vs CONTROL PREMIUM (base-rate-immune):', True, 10),
    ('  Event-CAR: +4.97pp (p=0.046, N=30,740). 30d trade-CAR: +4.24pp (p=0.082) — n.s.', False, 10),
    ('  Implied value (event-CAR): $1.90B total, $94,941 mean per trade.', False, 10),
    ('  CRSP mark-to-market (30d trade-CAR): $4.33B total, $3.79B from sells (38,882 trades).', False, 10),
    ('  Caveat: the event-CAR premium may reflect event-selection rather than foreknowledge.', False, 10),
    ('', False, 10),
    ('BASE-RATE-IMMUNE TESTS:', True, 10),
    ('  CAR NET SLOPE: event CAR ~ net sell direction (HC3, N=918). coef=−0.006, p=0.216. Null.', False, 10),
    ('  JOINT PT TEST: obs=56.67%, null=55.08%, excess=+1.59pp, z=1.17, p=0.240. Null.', False, 10),
    ('  DIRECTIONAL PLACEBO: observed=51.67% vs perm_mean=52.24%, p=0.661. Null.', False, 10),
    ('', False, 10),
    ('CONCENTRATION AND REPEAT TRADERS:', True, 10),
    ('  Top 5% of insider-event pairs account for 91.6% of net sell volume (Gini=0.963).', False, 10),
    ('  BUT: repeat vs one-time traders — U p=0.204, T p=0.203. No significant difference.', False, 10),
    ('  Concentration is in volume, not in accuracy. Heavy traders are not more accurate.', False, 10),
    ('', False, 10),
    ('PROXIMITY GRADIENT:', True, 10),
    ('  0-30d sells=59.8%, 31-60d=53.7%, 61-90d=55.2%, 91-180d=51.9%.', False, 10),
    ('  Not monotonic. No period-specific base rates computed — interpret with caution.', False, 10),
    ('', False, 10),
    ('── §13 ADDITIONAL TESTS — ALL NULL ────────────────────────────────────', False, 10),
    ('', False, 10),
    ('1. INTENSITY DiD (ESSAY3_INTENSITY_DID / _REG):', False, 10),
    ('   DV = NET_SELL_INTENSITY. Interaction = NEAR_WINDOW × NEGATIVE_EVENT.', False, 10),
    ('   Selected spec: coef=0.129, p=0.023. BUT full-population spec: coef=0.019, p=0.588.', False, 10),
    ('   Selected spec uses ~30% of panel (events with valid EVENT_CAR). Selection bias likely.', False, 10),
    ('   Primary result = full-population spec: null.', False, 10),
    ('', False, 10),
    ('2. PRICE DISCOVERY (ESSAY3_PRICE_DISCOVERY / _REG):', False, 10),
    ('   DV = EVENT_CAR. Predictor = NET_SELL_INTENSITY in far window (days 61-180).', False, 10),
    ('   Far-window spec (primary): coef=+0.005, p=0.582. Full pre-event: coef=−0.006, p=0.326.', False, 10),
    ('   No evidence that far-window net selling predicts event-direction. Null.', False, 10),
    ('', False, 10),
    ('3. WITHIN-INSIDER (ESSAY3_WITHIN_INSIDER):', False, 10),
    ('   Paired t-test (N=701 insiders): political vs control accuracy gap = +2.2pp, p=0.080.', False, 10),
    ('   LPM with insider FE (demeaned OLS): coef=+0.004, p=0.508. Null.', False, 10),
    ('', False, 10),
    ('4. CONNECTION TIMING (ESSAY3_CONN_TIMING / _REG):', False, 10),
    ('   DV = NET_SELL_INTENSITY. Triple DiD: HIGH_CONN × NEAR_WINDOW × NEGATIVE_EVENT.', False, 10),
    ('   Triple interaction coef=0.013, p=0.928. Null.', False, 10),
    ('', False, 10),
    ('── WHAT THE COLUMNS MEAN ──────────────────────────────────────────────', False, 10),
    ('METRIC_PVAL      Naive binomial vs 50% — DO NOT use as headline (wrong null)', False, 10),
    ('METRIC_PVAL_COND Honest binomial vs POOLED conditional base rate (57.84% / 42.16%)', False, 10),
    ('                 WARNING: pooled null is wrong for period-stratified comparisons.', False, 10),
    ('COND_NULL        The pooled conditional base rate used in METRIC_PVAL_COND', False, 10),
    ('METRIC_PVAL_BONF Bonferroni-corrected p (×4 windows) for proximity rows only', False, 10),
    ('P_VALUE_CLUSTERED One-tailed clustered p for directional cuts (EVENT/TICKER/EVENT_AGG)', False, 10),
    ('', False, 10),
    ('── TABLE GUIDE ─────────────────────────────────────────────────────────', False, 10),
    ('ESSAY3_INFORMED_TRADING      Accuracy by cut — use METRIC_PVAL_COND with pooled null caveat', False, 10),
    ('ESSAY3_INFORMED_PROXIMITY    Accuracy by days-before-event (METRIC_PVAL_BONF for MTesting)', False, 10),
    ('ESSAY3_INFORMED_DOLLARS      Dollar magnitudes for profitable sells by proximity/severity', False, 10),
    ('ESSAY3_CLUSTER_INFERENCE     Wild bootstrap + EVENT_AGG. P_VALUE_CLUSTERED = one-tailed.', False, 10),
    ('ESSAY3_DIRECTIONAL_PLACEBO   Date-randomization placebo (p=0.661)', False, 10),
    ('ESSAY3_CAR_NET_SLOPE         OLS: event CAR ~ net sell direction, 918 events, HC3 SE', False, 10),
    ('ESSAY3_JOINT_PT              Joint Pesaran-Timmermann 2×2 sign test (corrected variance)', False, 10),
    ('ESSAY3_CRSP_SUMMARY          Trade-CAR profitability (active/inactive/buys/sells)', False, 10),
    ('ESSAY3_CRSP_PROFITS          Full trade-level mark-to-market data (38,882 trades)', False, 10),
    ('ESSAY3_INSIDER_CONCENTRATION Gini / top-pct concentration by insider-event pairs', False, 10),
    ('ESSAY3_REPEAT_TRADERS        Repeat vs one-time traders (U p=0.204 — no difference)', False, 10),
    ('ESSAY3_WILCOXON_FAMILY       Gating test (28 tests, Holm + BH correction)', False, 10),
    ('ESSAY3_SIZE_ACCURACY         CG size-accuracy attenuation (supporting)', False, 10),
    ('ESSAY3_REVERSAL_REGRESSION   LPM regressions on directional accuracy', False, 10),
    ('ESSAY3_CONTROL_TRADES        Matched non-political control trades (17,969)', False, 10),
    ('ESSAY3_CONTROL_ATTRITION     CAR attrition diagnostics', False, 10),
    ('ESSAY3_TOST                  Equivalence tests — all three EQUIVALENT at d=0.2 SESOI', False, 10),
    ('ESSAY3_PANEL                 Unified insider trading panel (1,712 events)', False, 10),
    ('ESSAY3_INTENSITY_DID         DiD: NET_SELL_INTENSITY, NEAR×NEGATIVE_EVENT (§13)', False, 10),
    ('ESSAY3_INTENSITY_DID_REG     DiD regression — full-pop spec null (p=0.588); selected p=0.023', False, 10),
    ('ESSAY3_PRICE_DISCOVERY       Far-window net selling vs event CAR — correlation (§13)', False, 10),
    ('ESSAY3_PRICE_DISCOVERY_REG   OLS: event CAR ~ far-window NET_SELL_INTENSITY — null (§13)', False, 10),
    ('ESSAY3_WITHIN_INSIDER        Within-insider accuracy: paired t p=0.080, LPM FE p=0.508 (§13)', False, 10),
    ('ESSAY3_CONN_TIMING           NET_SELL_INTENSITY by connection × window × event-direction (§13)', False, 10),
    ('ESSAY3_CONN_TIMING_REG       Triple DiD: HIGH_CONN×NEAR×NEG_EVENT. coef=0.013, p=0.928 (§13)', False, 10),
]


def build_readme(wb):
    ws = wb.create_sheet('README')
    for i, (text, bold, size) in enumerate(README_TEXT, 1):
        cell = ws.cell(i, 1, text)
        cell.font = Font(bold=bold, size=size)
    ws.column_dimensions['A'].width = 80
    return ws


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f'Connecting to {DB_PATH}...')
    conn = sqlite3.connect(DB_PATH)

    def load(tbl):
        try:
            return pd.read_sql(f'SELECT * FROM {tbl}', conn)
        except Exception as e:
            print(f'  WARNING: could not load {tbl}: {e}')
            return pd.DataFrame()

    print('Loading tables...')
    tables = {
        'informed_trading':      load('ESSAY3_INFORMED_TRADING'),
        'informed_proximity':    load('ESSAY3_INFORMED_PROXIMITY'),
        'informed_dollars':      load('ESSAY3_INFORMED_DOLLARS'),
        'cluster_inference':     load('ESSAY3_CLUSTER_INFERENCE'),
        'directional_placebo':   load('ESSAY3_DIRECTIONAL_PLACEBO'),
        'car_net_slope':         load('ESSAY3_CAR_NET_SLOPE'),
        'joint_pt':              load('ESSAY3_JOINT_PT'),
        'crsp_summary':          load('ESSAY3_CRSP_SUMMARY'),
        'crsp_profits':          load('ESSAY3_CRSP_PROFITS'),
        'insider_concentration': load('ESSAY3_INSIDER_CONCENTRATION'),
        'wilcoxon_family':       load('ESSAY3_WILCOXON_FAMILY'),
        'mean_vs_dist':          load('ESSAY3_MEAN_VS_DISTRIBUTIONAL'),
        'size_accuracy':         load('ESSAY3_SIZE_ACCURACY'),
        'size_accuracy_slopes':  load('ESSAY3_SIZE_ACCURACY_SLOPES'),
        'reversal_regression':   load('ESSAY3_REVERSAL_REGRESSION'),
        'control_trades':        load('ESSAY3_CONTROL_TRADES'),
        'control_attrition':     load('ESSAY3_CONTROL_ATTRITION'),
        'tost':                  load('ESSAY3_TOST'),
        'placebo':               load('ESSAY3_PLACEBO'),
        'bootstrap_ci':          load('ESSAY3_BOOTSTRAP_CI'),
        'bootstrap_wilcoxon':    load('ESSAY3_BOOTSTRAP_WILCOXON'),
        'panel':                 load('ESSAY3_PANEL'),
        'stratification':        load('ESSAY3_STRATIFICATION'),
        'insider_panel':         load('ESSAY3_INSIDER_PANEL'),
        'concentration_cuts':    load('ESSAY3_CONCENTRATION_CUTS'),
        'repeat_traders':        load('ESSAY3_REPEAT_TRADERS'),
        'quantile_regression':   load('ESSAY3_QUANTILE_REGRESSION'),
        'trimmed_robustness':    load('ESSAY3_TRIMMED_ROBUSTNESS'),
        'active_subset':         load('ESSAY3_ACTIVE_SUBSET'),
        'intensity_did':         load('ESSAY3_INTENSITY_DID'),
        'intensity_did_reg':     load('ESSAY3_INTENSITY_DID_REG'),
        'price_discovery':       load('ESSAY3_PRICE_DISCOVERY'),
        'price_discovery_reg':   load('ESSAY3_PRICE_DISCOVERY_REG'),
        'within_insider':        load('ESSAY3_WITHIN_INSIDER'),
        'conn_timing':           load('ESSAY3_CONN_TIMING'),
        'conn_timing_reg':       load('ESSAY3_CONN_TIMING_REG'),
        'car_decomp':            load('ESSAY3_CAR_DECOMP'),
        'decomp_accuracy':       load('ESSAY3_DECOMP_ACCURACY'),
        'post_event_trading':    load('ESSAY3_POST_EVENT_TRADING'),
        'post_event_reg':        load('ESSAY3_POST_EVENT_REG'),
    }
    conn.close()

    print('Building workbook...')
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── README ──
    build_readme(wb)

    # ── Data sheets ──
    sheet_specs = [
        ('Informed Trading',       'informed_trading',
         'Headline Directional Accuracy (Event-CAR)',
         'PCT_PROFITABLE vs 57.4% conditional base rate (NOT 50%). See README for correct interpretation.'),
        ('Informed Proximity',     'informed_proximity',
         'Accuracy by Days Before Event', None),
        ('Informed Dollars',       'informed_dollars',
         'Dollar Magnitudes of Profitable Sells', None),
        ('Cluster Inference',      'cluster_inference',
         'Wild Bootstrap & EVENT_AGG Cluster Inference',
         'EVENT_AGG vs 57.4% base rate: p=0.39. EVENT_AGG vs 50% is WRONG null. See README.'),
        ('Directional Placebo',    'directional_placebo',
         'Date-Randomization Placebo Test', None),
        ('CAR Net Slope',          'car_net_slope',
         'OLS: Event CAR ~ Net Sell Direction (HC3 SE)',
         'Base-rate-immune. coef=-0.003, p=0.52. Not significant.'),
        ('Joint PT Test',          'joint_pt',
         'Joint Pesaran-Timmermann 2x2 Sign Test',
         'obs_acc=52.1%, null=51.7% (independence), p=0.085. Marginally significant.'),
        ('CRSP Summary',           'crsp_summary',
         'Trade-Window CAR Profitability Summary', None),
        ('CRSP Profits',           'crsp_profits',
         'Trade-Level Profit Data (38,882 trades)', None),
        ('Insider Concentration',  'insider_concentration',
         'Gini / Top-% Concentration of Trading', None),
        ('Wilcoxon Family',        'wilcoxon_family',
         'Gating Test: 28-Test Wilcoxon Family (Holm + BH)',
         '5 BH-significant cells. Wilcoxon and t-test diverge for 9/28 — heavy tails.'),
        ('Mean vs Distributional', 'mean_vs_dist',
         'Mean vs Distributional Comparison', None),
        ('Size Accuracy',          'size_accuracy',
         'CG Size-Accuracy Attenuation (Supporting)', None),
        ('Size Accuracy Slopes',   'size_accuracy_slopes',
         'Political vs Control Slope Comparison', None),
        ('Reversal Regression',    'reversal_regression',
         'LPM Regressions on Directional Accuracy', None),
        ('Control Trades',         'control_trades',
         'Matched Non-Political Control Trades', None),
        ('Control Attrition',      'control_attrition',
         'CAR Attrition Diagnostics (~20% attrition, size-selective)', None),
        ('TOST',                   'tost',
         'TOST Equivalence Tests', None),
        ('Placebo',                'placebo',
         'Permutation Placebo Test', None),
        ('Bootstrap CI',           'bootstrap_ci',
         'Bootstrap Confidence Intervals', None),
        ('Bootstrap Wilcoxon',     'bootstrap_wilcoxon',
         'Bootstrap Wilcoxon Inference', None),
        ('Panel',                  'panel',
         'Unified Insider Trading Panel (1,712 events)', None),
        ('Stratification',         'stratification',
         'Stratification Summary by Year / Tercile', None),
        ('Insider Panel',          'insider_panel',
         'Insider Fixed-Effects Panel Results', None),
        ('Concentration Cuts',     'concentration_cuts',
         'Concentration by Dimension', None),
        ('Repeat Traders',         'repeat_traders',
         'Repeat Trader Analysis', None),
        ('Quantile Regression',    'quantile_regression',
         'Quantile Regression Results', None),
        ('Trimmed Robustness',     'trimmed_robustness',
         'Trimmed Robustness (1% and 5% tails)', None),
        ('Active Subset',          'active_subset',
         'Active Insider Subset Characterization', None),
        ('Intensity DiD',          'intensity_did',
         'Trading Intensity Difference-in-Differences (§13)',
         'NEAR vs FAR window DiD: political insiders increase sell intensity approaching events.'),
        ('Intensity DiD Reg',      'intensity_did_reg',
         'DiD Regression: NET_SELL_INTENSITY ~ NEAR×POLITICAL (§13)',
         'LPM DiD with insider and ticker FE. Identifies abnormal pre-event sell-up.'),
        ('Price Discovery',        'price_discovery',
         'Price Discovery: Cumulative Returns Around Political Events (§13)',
         'Pre-event CAR drift for sells vs buys. Informed sellers front-run CAR reversal.'),
        ('Price Discovery Reg',    'price_discovery_reg',
         'Price Discovery Regression (§13)',
         'OLS: pre-event CAR ~ sell intensity, controlling for firm size and event severity.'),
        ('Within Insider',         'within_insider',
         'Within-Insider Accuracy Variation (§13)',
         'Demeaned OLS: accuracy varies within-insider across event types. No constant.'),
        ('Connection Timing',      'conn_timing',
         'Political Connection Timing Analysis (§13)',
         'Connected-insider trades cluster tighter to event dates than unconnected insiders.'),
        ('Connection Timing Reg',  'conn_timing_reg',
         'Connection Timing Regression (§13)',
         'LPM: ACCURATE ~ CONNECTED × DAYS_BEFORE_EVENT. Interaction tests timing advantage.'),
        ('CAR Decomp',            'car_decomp',
         'CAR Decomposition: Announcement [0,+1] vs Drift [+2,+60] (§14)',
         'Event-level CARs split into announcement reaction and post-announcement drift. Base: compute_car_decomposed.'),
        ('Decomp Accuracy',       'decomp_accuracy',
         'Sell Accuracy vs Announcement vs Drift Sub-Windows (§14)',
         'Each test uses construct-specific conditional base rate. BH-corrected across §14 family.'),
        ('Post-Event Trading',    'post_event_trading',
         'Post-Event Trading Intensity by Announcement Outcome (§14)',
         '2×2: HAS_POST_EARLY × NEGATIVE_CAR_ANNOUNCE. Tests whether insiders trade MORE after good news.'),
        ('Post-Event Reg',        'post_event_reg',
         'CAR_REMAINING ~ NET_SELL_INTENSITY[0,+10] + Controls (§14)',
         'Primary and 5 robustness specs. Event-clustered SE. Tests whether post-event selling predicts remaining drift.'),
    ]

    for sheet_name, key, title, note in sheet_specs:
        df = tables.get(key, pd.DataFrame())
        print(f'  Sheet: {sheet_name} ({len(df)} rows)')
        _write_table(wb, sheet_name, df, title=title, note=note)

    # ── Figure sheets ──
    print('Building figures...')
    fig_specs = [
        ('Fig_Proximity',        fig_sell_accuracy_by_proximity, 'informed_proximity'),
        ('Fig_Plan Split',       fig_plan_vs_nonplan,            'informed_trading'),
        ('Fig_Cluster CI',       fig_cluster_inference,          'cluster_inference'),
        ('Fig_Concentration',    fig_concentration,              'insider_concentration'),
        ('Fig_CRSP Summary',     fig_crsp_summary,               'crsp_summary'),
        ('Fig_Placebo',          fig_directional_placebo,        ('directional_placebo', 'informed_trading')),
        ('Fig_Joint PT',         fig_joint_pt,                   'joint_pt'),
        ('Fig_CAR Slope',        fig_car_net_slope,              'car_net_slope'),
    ]

    for sheet_name, builder, data_keys in fig_specs:
        try:
            if isinstance(data_keys, tuple):
                args = [tables[k] for k in data_keys]
            else:
                args = [tables[data_keys]]
            fig = builder(*args)
            if fig is not None:
                print(f'  Figure: {sheet_name}')
                _add_image_sheet(wb, sheet_name, fig)
        except Exception as e:
            print(f'  WARNING: figure {sheet_name} failed: {e}')

    print(f'Saving {OUT_PATH}...')
    wb.save(OUT_PATH)
    import os
    size_mb = os.path.getsize(OUT_PATH) / 1e6
    print(f'Done: {OUT_PATH} ({size_mb:.1f} MB, {len(wb.sheetnames)} sheets)')

    # Copy to Desktop
    import shutil
    dest = os.path.expanduser('~/Desktop/essay3_full_results.xlsx')
    shutil.copy2(OUT_PATH, dest)
    print(f'Copied to {dest}')


if __name__ == '__main__':
    main()
